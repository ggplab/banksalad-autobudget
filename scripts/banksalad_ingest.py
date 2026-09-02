#!/usr/bin/env python3
"""뱅크샐러드 엑셀 내보내기 → SQLite 적재.

뱅샐은 앱 전용이라 스크래핑이 불가능하다. 앱에서 "설정 → 데이터 내보내기 → 파일로 받기"를
실행하면 메일로 xlsx(zip)가 오고, 이 스크립트가 그걸 정규화해 누적 DB로 만든다.
내보내기는 최대 1년치라 월 1회면 11개월이 겹치므로 유실이 없다.

  python3 scripts/banksalad_ingest.py <xlsx경로> [...]
  python3 scripts/banksalad_ingest.py --stats

설계상 주의 3가지 (전부 실데이터에서 실측된 함정):

1. dedup 키에 발생순번이 필요하다.
   (날짜+시간+금액+내용+결제수단)만으로는 12개월 샘플에서 여러 행이 충돌한다.
   예: 지하철 요금 -3,100원이 같은 초에 3행 (정당한 별개 거래).
   시간이 00:00:00인 행도 많아 충돌 확률이 더 높다.

2. dedup이 아니라 upsert여야 한다.
   사용자가 뱅샐 앱에서 과거 건을 재분류하면 다음 export에 반영돼 오는데,
   "이미 있음"으로 버리면 그 수정이 영원히 반영되지 않는다.
   단 최신 export에 없는 기존 행을 삭제하지는 않는다 — 계좌 연동을 해지하면
   과거 내역이 export에서 빠질 수 있고, 그때 DB가 지워지면 복구 불가다.

3. 환불은 '지출' 타입의 양수 행이다.
   합계를 abs로 낼지 signed로 낼지 정의하지 않으면 리포트 검증이 불가능하다.
   이 스크립트는 원본 부호를 그대로 보존하고, 판단은 소비 측에 맡긴다.
"""
import hashlib
import os
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DB_PATH = Path(os.environ.get("BUDGET_DB") or REPO / "data" / "budget.db")
SHEET = "가계부 내역"
COLUMNS = ["날짜", "시간", "타입", "대분류", "소분류", "내용", "금액", "화폐", "결제수단", "메모"]

SCHEMA = """
CREATE TABLE IF NOT EXISTS transactions (
    id           TEXT PRIMARY KEY,   -- sha256(자연키)#발생순번
    date         TEXT NOT NULL,      -- YYYY-MM-DD
    time         TEXT NOT NULL,      -- HH:MM:SS
    type         TEXT NOT NULL,      -- 지출 | 수입 | 이체
    category     TEXT,               -- 대분류
    subcategory  TEXT,               -- 소분류
    content      TEXT,               -- 가맹점/내용
    amount       INTEGER NOT NULL,   -- 원본 부호 보존 (지출 양수 = 환불)
    currency     TEXT,
    method       TEXT,               -- 결제수단
    memo         TEXT,
    source_file  TEXT NOT NULL,      -- 마지막으로 이 행을 갱신한 export
    first_seen   TEXT NOT NULL,
    last_seen    TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_tx_date   ON transactions(date);
CREATE INDEX IF NOT EXISTS idx_tx_type   ON transactions(type);
CREATE INDEX IF NOT EXISTS idx_tx_method ON transactions(method);

CREATE TABLE IF NOT EXISTS ingest_log (
    source_file TEXT NOT NULL,
    ingested_at TEXT NOT NULL,
    rows_total  INTEGER NOT NULL,
    rows_new    INTEGER NOT NULL,
    rows_updated INTEGER NOT NULL
);
"""


def natural_key(r: dict) -> str:
    raw = f"{r['date']}|{r['time']}|{r['amount']}|{r['content']}|{r['method']}"
    return hashlib.sha256(raw.encode()).hexdigest()[:24]


def read_rows(path: Path) -> list[dict]:
    """xlsx의 '가계부 내역' 시트를 dict 리스트로. 발생순번까지 부여한다."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"'{SHEET}' 시트 없음: {path.name} (시트: {wb.sheetnames})")
    ws = wb[SHEET]

    it = ws.iter_rows(values_only=True)
    header = next(it)
    if list(header)[:len(COLUMNS)] != COLUMNS:
        sys.exit(f"컬럼 구조가 다름: {path.name}\n  기대: {COLUMNS}\n  실제: {list(header)}")

    rows: list[dict] = []
    seen: defaultdict[str, int] = defaultdict(int)
    for raw in it:
        if not raw or raw[0] is None:
            continue
        date = str(raw[0])[:10]
        rec = {
            "date": date,
            "time": str(raw[1] or "00:00:00"),
            "type": str(raw[2] or ""),
            "category": str(raw[3] or ""),
            "subcategory": str(raw[4] or ""),
            "content": str(raw[5] or ""),
            "amount": int(raw[6] or 0),
            "currency": str(raw[7] or "KRW"),
            "method": str(raw[8] or ""),
            "memo": str(raw[9] or "") if len(raw) > 9 and raw[9] else "",
        }
        k = natural_key(rec)
        seen[k] += 1
        rec["id"] = f"{k}#{seen[k]}"     # 같은 초 동일거래 다건을 보존
        rows.append(rec)
    wb.close()
    return rows


def ingest(path: Path, conn: sqlite3.Connection) -> tuple[int, int, int]:
    rows = read_rows(path)
    now = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    src = path.name
    new = updated = 0

    for r in rows:
        cur = conn.execute("SELECT category, subcategory, memo FROM transactions WHERE id = ?",
                           (r["id"],)).fetchone()
        if cur is None:
            conn.execute(
                """INSERT INTO transactions
                   (id,date,time,type,category,subcategory,content,amount,currency,method,memo,
                    source_file,first_seen,last_seen)
                   VALUES (:id,:date,:time,:type,:category,:subcategory,:content,:amount,
                           :currency,:method,:memo,:src,:now,:now)""",
                {**r, "src": src, "now": now})
            new += 1
        else:
            # 최신 export가 이긴다 — 앱에서 재분류한 결과를 반영하기 위해.
            # 없어진 행을 지우지는 않는다(연동 해지 시 과거 내역 유실 방지).
            changed = (cur[0], cur[1], cur[2]) != (r["category"], r["subcategory"], r["memo"])
            conn.execute(
                """UPDATE transactions
                   SET category=:category, subcategory=:subcategory, memo=:memo,
                       source_file=:src, last_seen=:now
                   WHERE id=:id""",
                {**r, "src": src, "now": now})
            if changed:
                updated += 1

    conn.execute(
        "INSERT INTO ingest_log (source_file,ingested_at,rows_total,rows_new,rows_updated) "
        "VALUES (?,?,?,?,?)", (src, now, len(rows), new, updated))
    conn.commit()
    return len(rows), new, updated


def stats(conn: sqlite3.Connection) -> None:
    q = lambda s, *a: conn.execute(s, a).fetchone()  # noqa: E731
    total = q("SELECT COUNT(*) FROM transactions")[0]
    print(f"거래 {total:,}건")
    lo, hi = q("SELECT MIN(date), MAX(date) FROM transactions")
    print(f"기간 {lo} ~ {hi}")

    print("\n타입별:")
    for t, c, s in conn.execute(
            "SELECT type, COUNT(*), SUM(amount) FROM transactions GROUP BY type ORDER BY COUNT(*) DESC"):
        print(f"  {t:5} {c:6,}건  signed {s:>15,}원")

    refund = q("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM transactions WHERE type='지출' AND amount > 0")
    print(f"\n환불(지출 타입 양수): {refund[0]}건, +{refund[1]:,}원")

    dup = conn.execute(
        "SELECT COUNT(*) FROM (SELECT substr(id,1,24) k FROM transactions "
        "GROUP BY k HAVING COUNT(*) > 1)").fetchone()[0]
    extra = q("SELECT COUNT(*) FROM transactions WHERE id NOT LIKE '%#1'")[0]
    print(f"동일 자연키 다건 그룹: {dup}개 (발생순번으로 보존된 추가 행 {extra}개)")

    print("\n적재 이력:")
    for r in conn.execute("SELECT source_file, ingested_at, rows_total, rows_new, rows_updated "
                          "FROM ingest_log ORDER BY ingested_at"):
        print(f"  {r[1][:19]}  {r[0][:44]:46} 총{r[2]:5} 신규{r[3]:5} 갱신{r[4]:4}")


def main() -> None:
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(1)

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)

    if args[0] == "--stats":
        stats(conn)
        return

    for a in args:
        p = Path(a).expanduser().resolve()
        if not p.exists():
            sys.exit(f"파일 없음: {p}")
        total, new, upd = ingest(p, conn)
        print(f"{p.name}: 총 {total:,}행 → 신규 {new:,} / 갱신 {upd:,}")
    conn.close()


if __name__ == "__main__":
    main()
